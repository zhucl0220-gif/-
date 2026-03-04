"""
app/tools/statistics_tools.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
统计报表 Agent Tools (AD-18 ~ AD-21)

对外暴露：
  get_dashboard_statistics()
      计算全局仪表盘统计：患者阶段分布、风险等级分布、
      近 30 天 Agent 功能使用趋势、核心 KPI 汇总。

  export_system_data(target, fmt)
      导出系统数据为 Excel 文件，返回临时下载 URL。
      target: 'patients' | 'lab_results' | 'diet_records' | 'alerts'
      fmt:    'excel'（预留 'csv' 扩展）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
from __future__ import annotations

import logging
import os
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import func, select, and_

from app.config import settings
from app.database import AsyncSessionLocal
from app.models.models import (
    AgentTask, AgentTaskStatus, AgentTaskType,
    AlertSeverity, AlertStatus,
    DietRecord, LabResult, NutritionPlan,
    PatientProfile, RiskAlert, TransplantPhase,
)

logger = logging.getLogger(__name__)

# ── 导出目录 ────────────────────────────────────────────────────────────────
_EXPORT_DIR = Path(settings.UPLOAD_DIR) / "exports"
_EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# ── 移植阶段中文标签 ─────────────────────────────────────────────────────────
_PHASE_LABELS: dict[str, str] = {
    TransplantPhase.PRE_ASSESSMENT.value:   "术前评估期",
    TransplantPhase.PRE_OPERATION.value:    "术前准备期",
    TransplantPhase.EARLY_POST_OP.value:    "术后早期",
    TransplantPhase.RECOVERY.value:         "恢复期",
    TransplantPhase.REHABILITATION.value:   "康复期",
    TransplantPhase.LONG_TERM_FOLLOW.value: "长期随访",
}

# ── Agent 任务类型中文标签 ───────────────────────────────────────────────────
_TASK_LABELS: dict[str, str] = {
    AgentTaskType.LAB_ANALYSIS.value:    "检验单解读",
    AgentTaskType.NUTRITION_PLAN.value:  "营养方案生成",
    AgentTaskType.WEB_SEARCH.value:      "网络搜索",
    AgentTaskType.CODE_EXECUTION.value:  "代码沙箱",
    AgentTaskType.DIET_EVALUATION.value: "饮食评估",
    AgentTaskType.GENERAL_QA.value:      "通用问答",
}


# ══════════════════════════════════════════════════════════════════════════════
# Tool 1 – 仪表盘统计数据
# ══════════════════════════════════════════════════════════════════════════════

async def get_dashboard_statistics() -> dict[str, Any]:
    """
    计算全局仪表盘统计，供前端图表直接消费。

    返回结构：
    {
      "kpi": {
        "total_patients": int,
        "active_alerts": int,
        "avg_compliance": float | null,   # 近30天饮食依从性均分
        "total_agent_tasks_30d": int,
      },
      "phase_distribution": [
        {"phase": "early_post_op", "label": "术后早期", "count": 12}, ...
      ],
      "risk_distribution": [
        {"level": "critical", "label": "危急", "count": 3}, ...
      ],
      "usage_trend_30d": [
        {"date": "2026-02-01", "lab_analysis": 2, "nutrition_plan": 1, ...}, ...
      ],
    }
    """
    async with AsyncSessionLocal() as session:
        today = date.today()
        thirty_days_ago = today - timedelta(days=30)

        # ── KPI ───────────────────────────────────────────────────────────────
        total_patients = (
            await session.execute(select(func.count()).select_from(PatientProfile))
        ).scalar_one()

        active_alerts = (
            await session.execute(
                select(func.count()).select_from(RiskAlert)
                .where(RiskAlert.status == AlertStatus.ACTIVE)
            )
        ).scalar_one()

        # 近 30 天饮食打卡依从性均分
        avg_comp_row = (
            await session.execute(
                select(func.avg(DietRecord.compliance_score))
                .where(
                    and_(
                        DietRecord.compliance_score.is_not(None),
                        DietRecord.record_date >= thirty_days_ago,
                    )
                )
            )
        ).scalar_one()
        avg_compliance = round(float(avg_comp_row), 1) if avg_comp_row is not None else None

        total_tasks_30d = (
            await session.execute(
                select(func.count()).select_from(AgentTask)
                .where(AgentTask.created_at >= datetime.combine(thirty_days_ago, datetime.min.time()))
            )
        ).scalar_one()

        # ── 阶段分布 ─────────────────────────────────────────────────────────
        phase_rows = (
            await session.execute(
                select(PatientProfile.current_phase, func.count().label("cnt"))
                .group_by(PatientProfile.current_phase)
            )
        ).all()

        # 保证所有阶段都有条目（即便 count=0）
        phase_map: dict[str, int] = defaultdict(int)
        for row in phase_rows:
            phase_map[row.current_phase.value if hasattr(row.current_phase, 'value') else str(row.current_phase)] += row.cnt

        phase_distribution = [
            {
                "phase": phase.value,
                "label": _PHASE_LABELS.get(phase.value, phase.value),
                "count": phase_map.get(phase.value, 0),
            }
            for phase in TransplantPhase
        ]

        # ── 风险等级分布 ─────────────────────────────────────────────────────
        # 从 RiskAlert 推导每位患者的最高风险等级
        alert_rows = (
            await session.execute(
                select(
                    RiskAlert.patient_id,
                    RiskAlert.severity,
                    RiskAlert.status,
                )
            )
        ).all()

        # 构建 patient_id → risk_level 映射
        patient_risk: dict[Any, str] = {}
        for r in alert_rows:
            pid = str(r.patient_id)
            sev = r.severity.value if hasattr(r.severity, 'value') else str(r.severity)
            st  = r.status.value  if hasattr(r.status, 'value')    else str(r.status)
            cur = patient_risk.get(pid)
            if st == AlertStatus.ACTIVE.value:
                if sev == AlertSeverity.CRITICAL.value:
                    patient_risk[pid] = "critical"
                elif sev == AlertSeverity.WARNING.value and cur != "critical":
                    patient_risk[pid] = "high"
                elif sev == AlertSeverity.INFO.value and cur not in ("critical", "high"):
                    patient_risk[pid] = "medium"
            elif st == AlertStatus.ACKNOWLEDGED.value and cur is None:
                patient_risk[pid] = "low"

        # 统计全部患者
        all_patient_ids = (
            await session.execute(select(PatientProfile.id))
        ).scalars().all()

        risk_counter: dict[str, int] = defaultdict(int)
        for pid in all_patient_ids:
            level = patient_risk.get(str(pid), "safe")
            risk_counter[level] += 1

        risk_distribution = [
            {"level": "critical", "label": "危急",   "count": risk_counter.get("critical", 0), "color": "#FF4D4F"},
            {"level": "high",     "label": "高风险",  "count": risk_counter.get("high",     0), "color": "#FA8C16"},
            {"level": "medium",   "label": "中风险",  "count": risk_counter.get("medium",   0), "color": "#FADB14"},
            {"level": "safe",     "label": "低风险",  "count": risk_counter.get("safe",     0), "color": "#52C41A"},
        ]

        # ── 近 30 天 Agent 功能使用趋势 ─────────────────────────────────────
        task_rows = (
            await session.execute(
                select(
                    func.date(AgentTask.created_at).label("day"),
                    AgentTask.task_type,
                    func.count().label("cnt"),
                )
                .where(AgentTask.created_at >= datetime.combine(thirty_days_ago, datetime.min.time()))
                .group_by(func.date(AgentTask.created_at), AgentTask.task_type)
                .order_by(func.date(AgentTask.created_at))
            )
        ).all()

        # 构建日期 → {task_type: count} 映射
        trend_map: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for row in task_rows:
            day_str = str(row.day)
            tt = row.task_type.value if hasattr(row.task_type, 'value') else str(row.task_type)
            trend_map[day_str][tt] = row.cnt

        # 填充连续30天，缺失日期补0
        usage_trend: list[dict[str, Any]] = []
        for i in range(30):
            d = (thirty_days_ago + timedelta(days=i)).isoformat()
            entry: dict[str, Any] = {"date": d}
            for tt in AgentTaskType:
                entry[tt.value] = trend_map.get(d, {}).get(tt.value, 0)
            usage_trend.append(entry)

        return {
            "kpi": {
                "total_patients":       total_patients,
                "active_alerts":        active_alerts,
                "avg_compliance":       avg_compliance,
                "total_agent_tasks_30d": total_tasks_30d,
            },
            "phase_distribution":  phase_distribution,
            "risk_distribution":   risk_distribution,
            "usage_trend_30d":     usage_trend,
            "task_labels":         _TASK_LABELS,
        }


# ══════════════════════════════════════════════════════════════════════════════
# Tool 2 – 系统数据导出（Agent-callable）
# ══════════════════════════════════════════════════════════════════════════════

async def export_system_data(
    target: str = "patients",
    fmt: str = "excel",
) -> dict[str, Any]:
    """
    导出系统数据为 Excel，写入临时文件并返回可下载 URL。

    target: 'patients' | 'lab_results' | 'diet_records' | 'alerts'
    fmt:    'excel'（默认）

    返回:
    {
      "success": True,
      "filename": "export_patients_20260303_xxxxxxxx.xlsx",
      "download_url": "/api/v1/statistics/download/export_patients_...",
      "row_count": 42,
      "message": "...",
    }
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"success": False, "error": "openpyxl 未安装，请运行 pip install openpyxl"}

    async with AsyncSessionLocal() as session:
        wb = openpyxl.Workbook()
        ws = wb.active

        # ── 通用样式 ──────────────────────────────────────────────────────────
        header_font   = Font(bold=True, color="FFFFFF", size=11)
        header_fill   = PatternFill("solid", fgColor="1677FF")
        header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center_align  = Alignment(horizontal="center", vertical="center")

        def set_header(ws, headers: list[str]) -> None:
            ws.row_dimensions[1].height = 28
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font   = header_font
                cell.fill   = header_fill
                cell.alignment = header_align

        def auto_width(ws) -> None:
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── 患者全景 ──────────────────────────────────────────────────────────
        if target == "patients":
            ws.title = "患者档案"
            headers = [
                "姓名", "性别", "出生日期", "手机", "身高(cm)", "体重(kg)", "BMI",
                "诊断", "移植日期", "当前阶段", "建档时间",
            ]
            set_header(ws, headers)
            rows_q = (await session.execute(
                select(PatientProfile).order_by(PatientProfile.created_at.desc())
            )).scalars().all()

            for i, p in enumerate(rows_q, 2):
                gender_map = {"male": "男", "female": "女", "unknown": "未知"}
                ws.append([
                    p.name,
                    gender_map.get(p.gender.value if hasattr(p.gender, 'value') else str(p.gender), "—"),
                    str(p.birth_date) if p.birth_date else "",
                    p.phone or "",
                    p.height_cm or "",
                    p.weight_kg or "",
                    round(p.bmi, 1) if p.bmi else "",
                    p.diagnosis or "",
                    str(p.transplant_date) if p.transplant_date else "",
                    _PHASE_LABELS.get(p.current_phase.value if hasattr(p.current_phase, 'value') else str(p.current_phase), ""),
                    p.created_at.strftime("%Y-%m-%d") if p.created_at else "",
                ])
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).alignment = center_align

        # ── 检验结果 ──────────────────────────────────────────────────────────
        elif target == "lab_results":
            ws.title = "检验结果"
            headers = ["患者姓名", "报告日期", "报告类型", "AI 解读摘要", "是否已分析", "录入时间"]
            set_header(ws, headers)
            rows_q = (await session.execute(
                select(LabResult, PatientProfile.name)
                .join(PatientProfile, LabResult.patient_id == PatientProfile.id)
                .order_by(LabResult.report_date.desc())
            )).all()

            for i, (lr, pname) in enumerate(rows_q, 2):
                ai_sum = ""
                if lr.analysis_result and isinstance(lr.analysis_result, dict):
                    ai_sum = lr.analysis_result.get("summary", "")[:200]
                ws.append([
                    pname,
                    str(lr.report_date) if lr.report_date else "",
                    lr.report_type or "",
                    ai_sum,
                    "是" if lr.is_analyzed else "否",
                    lr.created_at.strftime("%Y-%m-%d") if lr.created_at else "",
                ])
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).alignment = center_align

        # ── 饮食打卡 ──────────────────────────────────────────────────────────
        elif target == "diet_records":
            ws.title = "饮食打卡"
            headers = ["患者姓名", "打卡日期", "餐次", "合计热量(kcal)", "合计蛋白(g)", "依从性评分"]
            set_header(ws, headers)
            rows_q = (await session.execute(
                select(DietRecord, PatientProfile.name)
                .join(PatientProfile, DietRecord.patient_id == PatientProfile.id)
                .order_by(DietRecord.record_date.desc())
            )).all()

            for i, (dr, pname) in enumerate(rows_q, 2):
                ws.append([
                    pname,
                    str(dr.record_date),
                    dr.meal_type,
                    round(dr.total_calories, 1) if dr.total_calories else "",
                    round(dr.total_protein_g, 1) if dr.total_protein_g else "",
                    round(dr.compliance_score, 1) if dr.compliance_score else "",
                ])
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).alignment = center_align

        # ── 风险预警 ──────────────────────────────────────────────────────────
        elif target == "alerts":
            ws.title = "风险预警"
            headers = ["患者姓名", "预警类型", "严重程度", "状态", "指标名称", "实测值", "阈值", "单位", "预警信息", "创建时间", "处理医生", "处理备注"]
            set_header(ws, headers)
            rows_q = (await session.execute(
                select(RiskAlert, PatientProfile.name)
                .join(PatientProfile, RiskAlert.patient_id == PatientProfile.id)
                .order_by(RiskAlert.created_at.desc())
            )).all()

            type_labels  = {"abnormal_indicator": "指标异常", "high_risk_patient": "高风险患者", "weight_loss": "体重骤降", "low_compliance": "依从性低"}
            sev_labels   = {"critical": "危急", "warning": "警告", "info": "提示"}
            status_labels = {"active": "待处理", "acknowledged": "已确认", "resolved": "已解除"}

            for i, (ra, pname) in enumerate(rows_q, 2):
                at   = ra.alert_type.value  if hasattr(ra.alert_type, 'value')  else str(ra.alert_type)
                sev  = ra.severity.value    if hasattr(ra.severity, 'value')    else str(ra.severity)
                st   = ra.status.value      if hasattr(ra.status, 'value')      else str(ra.status)
                ws.append([
                    pname,
                    type_labels.get(at, at),
                    sev_labels.get(sev, sev),
                    status_labels.get(st, st),
                    ra.metric_name   or "",
                    ra.metric_value  or "",
                    ra.threshold_value or "",
                    ra.unit          or "",
                    ra.message,
                    ra.created_at.strftime("%Y-%m-%d %H:%M") if ra.created_at else "",
                    ra.acknowledged_by or "",
                    ra.resolve_note  or "",
                ])
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).alignment = center_align

        else:
            return {"success": False, "error": f"不支持的导出目标: {target}"}

        auto_width(ws)
        row_count = ws.max_row - 1  # 减去表头行

        # 生成唯一文件名并保存
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        uid = uuid.uuid4().hex[:8]
        filename = f"export_{target}_{ts}_{uid}.xlsx"
        filepath = _EXPORT_DIR / filename
        wb.save(str(filepath))

        download_url = f"/api/v1/statistics/download/{filename}"
        logger.info("导出完成: %s (%d 行)", filename, row_count)

        return {
            "success":      True,
            "filename":     filename,
            "download_url": download_url,
            "row_count":    row_count,
            "message":      f"已成功导出 {row_count} 条 {target} 数据",
        }
